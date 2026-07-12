#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
------------------------------------------------------------
GSPL-01_Rhino_Extractor.py

GIAR Simulation Pipeline for LiDAR

Version : 0.2

Author:
    Ing. Pablo Daniel Folino

Description:
    Opens a Rhino (.3dm) model and inspects its contents.

------------------------------------------------------------
"""

from pathlib import Path
import json
import logging
import sys
import rhino3dm
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation


CONFIG_FILE = "config.json"


# ==========================================================
# Rhino Extractor
# ==========================================================

class RhinoExtractor:

    def __init__(self):

        self.cfg = None
        self.logger = None
        self.model = None


    # ------------------------------------------------------
    # Configuration
    # ------------------------------------------------------
    def load_config(self):

        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            self.cfg = json.load(f)


    # ------------------------------------------------------
    # Logger
    # ------------------------------------------------------
    def configure_logger(self):

        log_dir = Path(self.cfg["paths"]["log_directory"])

        log_dir.mkdir(parents=True, exist_ok=True)

        logfile = log_dir / "GSPL-01.log"

        logging.basicConfig(

            level=logging.INFO,

            format="%(asctime)s | %(levelname)-8s | %(message)s",

            handlers=[
                logging.FileHandler(logfile, mode="w"),
                logging.StreamHandler(sys.stdout)
            ]
        )

        self.logger = logging.getLogger("GSPL")


    # ------------------------------------------------------
    # Banner
    # ------------------------------------------------------
    def print_banner(self):

        self.logger.info("")
        self.logger.info("====================================================")
        self.logger.info(" GSPL-01 Rhino Extractor")
        self.logger.info(" Version : 0.2")
        self.logger.info("====================================================")
        self.logger.info("")

        self.logger.info("Project : %s",
                         self.cfg["project"]["name"])

        self.logger.info("Version : %s",
                         self.cfg["project"]["version"])

        self.logger.info("")


    # ------------------------------------------------------
    # Verify project
    # ------------------------------------------------------
    def verify_project(self):

        for directory in [

            self.cfg["paths"]["input_directory"],
            self.cfg["paths"]["stl_directory"],
            self.cfg["paths"]["database_directory"],
            self.cfg["paths"]["output_directory"]

        ]:

            if Path(directory).exists():

                self.logger.info("[ OK ] %s", directory)

            else:

                self.logger.error("[ERROR] %s", directory)


    # ------------------------------------------------------
    # Open Rhino Model
    # ------------------------------------------------------
    def open_model(self):

        filename = (

            Path(self.cfg["paths"]["input_directory"])

            /

            self.cfg["files"]["input_file_3dm"]

        )

        self.logger.info("")
        self.logger.info("Opening Rhino model...")
        self.logger.info("%s", filename)

        self.model = rhino3dm.File3dm.Read(str(filename))

        if self.model is None:

            raise RuntimeError("Unable to open Rhino file.")

        self.logger.info("[ OK ] Rhino model loaded.")


    # ------------------------------------------------------
    # Model information
    # ------------------------------------------------------
    def print_model_information(self):

        self.logger.info("")
        self.logger.info("----------------------------------------")
        self.logger.info("Model Information")
        self.logger.info("----------------------------------------")

        self.logger.info("File Version : %s",
                         self.model.Settings.ModelUnitSystem)

        self.logger.info("Layers       : %d",
                         len(self.model.Layers))

        self.logger.info("Objects      : %d",
                         len(self.model.Objects))

        self.logger.info("Groups       : %d",
                         len(self.model.Groups))


    # ------------------------------------------------------
    # Inspect Objects
    # ------------------------------------------------------
    def inspect_objects(self):

        self.logger.info("")
        self.logger.info("----------------------------------------")
        self.logger.info("Objects")
        self.logger.info("----------------------------------------")

        for i, obj in enumerate(self.model.Objects, start=1):

            att = obj.Attributes

            geo = obj.Geometry

            self.logger.info("")
            self.logger.info("[%04d]", i)

            self.logger.info("Name     : %s", att.Name)

            self.logger.info("GUID     : %s", att.Id)

            self.logger.info("Layer Id : %d", att.LayerIndex)

            self.logger.info("Geometry : %s",
                             geo.ObjectType)

    # ------------------------------------------------------
    # Geometry Audit
    # ------------------------------------------------------
    def audit_geometry(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" GEOMETRY AUDIT")
        self.logger.info("========================================================")

        errors = 0

        for obj in self.model.Objects:

            att = obj.Attributes
            geo = obj.Geometry

            name = att.Name if att.Name else "<UNNAMED>"

            # Obtener nombre de la Layer
            try:
                layer = self.model.Layers[att.LayerIndex].Name
            except:
                layer = "<UNKNOWN>"

            guid = att.Id

            # --------------------------------------------------
            # BREP
            # --------------------------------------------------

            if isinstance(geo, rhino3dm.Brep):

                if geo.IsSolid:

                    continue

                errors += 1

                self.logger.error("")
                self.logger.error("ERROR %03d", errors)
                self.logger.error("--------------------------------------------")
                self.logger.error("Component : %s", name)
                self.logger.error("Layer     : %s", layer)
                self.logger.error("GUID      : %s", guid)
                self.logger.error("Geometry  : Open Brep")
                self.logger.error("")
                self.logger.error("Action:")
                self.logger.error("    Convert this Brep into a CLOSED solid.")
                self.logger.error("--------------------------------------------")

                continue

            # --------------------------------------------------
            # EXTRUSION
            # --------------------------------------------------

            if isinstance(geo, rhino3dm.Extrusion):

                if geo.IsSolid:

                    continue

                errors += 1

                self.logger.error("")
                self.logger.error("ERROR %03d", errors)
                self.logger.error("--------------------------------------------")
                self.logger.error("Component : %s", name)
                self.logger.error("Layer     : %s", layer)
                self.logger.error("GUID      : %s", guid)
                self.logger.error("Geometry  : Open Extrusion")
                self.logger.error("")
                self.logger.error("Action:")
                self.logger.error("    Close the extrusion.")
                self.logger.error("--------------------------------------------")

                continue

            # --------------------------------------------------
            # INVALID GEOMETRY
            # --------------------------------------------------

            errors += 1

            self.logger.error("")
            self.logger.error("ERROR %03d", errors)
            self.logger.error("--------------------------------------------")
            self.logger.error("Component : %s", name)
            self.logger.error("Layer     : %s", layer)
            self.logger.error("GUID      : %s", guid)
            self.logger.error("Geometry  : %s", geo.ObjectType)
            self.logger.error("")
            self.logger.error("Action:")
            self.logger.error("    Delete or convert this object into a closed solid.")
            self.logger.error("--------------------------------------------")

        self.logger.info("")
        self.logger.info("========================================================")

        if errors == 0:

            self.logger.info("Geometry Audit : PASSED")

        else:

            self.logger.error("Geometry Audit : FAILED")
            self.logger.error("Total Errors   : %d", errors)

            raise RuntimeError(
                f"Geometry Audit FAILED ({errors} errors)"
            )

        self.logger.info("========================================================")

    # ------------------------------------------------------
    # Initialize Database
    # ------------------------------------------------------
    def initialize_database(self):

        self.database = {

            "project": {

                "name": self.cfg["project"]["name"],
                "version": self.cfg["project"]["version"]

            },

            "statistics": {

                "components": 0,
                "entities": 0

            },

            "components": []

        }
    
    # ------------------------------------------------------
    # Build Component Database
    # ------------------------------------------------------
    def build_database(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Building Component Database")
        self.logger.info("========================================================")

        self.initialize_database()

        components = {}

        for index, obj in enumerate(self.model.Objects):

            att = obj.Attributes
            geo = obj.Geometry

            # --------------------------------------------------
            # Component Name
            # --------------------------------------------------

            name = att.Name.strip() if att.Name else "<UNNAMED>"

            # --------------------------------------------------
            # Layer
            # --------------------------------------------------

            try:
                layer = self.model.Layers[att.LayerIndex].Name
            except Exception:
                layer = "<UNKNOWN>"

            # --------------------------------------------------
            # Geometry Type
            # --------------------------------------------------

            geometry = str(geo.ObjectType).replace("ObjectType.", "")

            # --------------------------------------------------
            # Create Component
            # --------------------------------------------------

            if name not in components:

                component = {

                    "id": len(components) + 1,

                    "uuid": str(uuid.uuid4()),

                    "name": name,

                    "type": "component",

                    "layer": layer,

                    "parent": None,

                    "children": [],

                    "stl": None,

                    "transform": None,

                    "bounding_box": None,

                    "center": None,

                    "mass": None,

                    "material": None,

                    "properties": {},

                    "status": {

                        "geometry_audit": "PASSED",
                        "hierarchy_audit": "PENDING",
                        "stl_export": "PENDING",
                        "simulation": "PENDING"

                    },

                    "entities": []

                }

                components[name] = component

            # --------------------------------------------------
            # Add Entity
            # --------------------------------------------------

            entity = {

                "index": index,

                "guid": str(att.Id),

                "geometry": geometry

            }

            components[name]["entities"].append(entity)

        # ------------------------------------------------------
        # Sort Components
        # ------------------------------------------------------

        component_list = sorted(
            components.values(),
            key=lambda c: c["name"].lower()
        )

        # ------------------------------------------------------
        # Reassign IDs after sorting
        # ------------------------------------------------------

        for new_id, component in enumerate(component_list, start=1):

            component["id"] = new_id

        # ------------------------------------------------------
        # Statistics
        # ------------------------------------------------------

        self.database["statistics"]["components"] = len(component_list)

        self.database["statistics"]["entities"] = len(self.model.Objects)

        self.database["components"] = component_list

        # ------------------------------------------------------
        # Log
        # ------------------------------------------------------

        self.logger.info("Components : %d", len(component_list))
        self.logger.info("Entities   : %d", len(self.model.Objects))

        self.logger.info("")
        self.logger.info("Component List")
        self.logger.info("--------------------------------------------------------")

        for component in component_list:

            self.logger.info(
                "[%03d] %-35s (%3d entities)",
                component["id"],
                component["name"],
                len(component["entities"])
            )

        self.logger.info("--------------------------------------------------------")

    # ------------------------------------------------------
    # Save Database
    # ------------------------------------------------------
    def save_database(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Saving Component Database")
        self.logger.info("========================================================")

        # --------------------------------------------------
        # Output directory
        # --------------------------------------------------

        database_dir = Path(self.cfg["paths"]["database_directory"])

        database_dir.mkdir(parents=True, exist_ok=True)

        # --------------------------------------------------
        # Output file
        # --------------------------------------------------

        output_file = (
            database_dir /
            self.cfg["files"]["model_database"]
        )

        # --------------------------------------------------
        # Save JSON
        # --------------------------------------------------

        try:

            with open(output_file, "w", encoding="utf-8") as f:

                json.dump(
                    self.database,
                    f,
                    indent=4,
                    ensure_ascii=False
                )

            self.logger.info("")
            self.logger.info("[ OK ] Database successfully saved.")
            self.logger.info("File       : %s", output_file)
            self.logger.info("Components : %d",
                            self.database["statistics"]["components"])
            self.logger.info("Entities   : %d",
                            self.database["statistics"]["entities"])

        except Exception as e:

            self.logger.error("")
            self.logger.error("[ERROR] Unable to save database.")
            self.logger.error(str(e))
            raise

        self.logger.info("========================================================")

    # ------------------------------------------------------
    # Generate Assembly Template
    # ------------------------------------------------------
    def generate_assembly_template(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Generating Assembly Template")
        self.logger.info("========================================================")

        # --------------------------------------------------
        # Create Assembly Structure
        # --------------------------------------------------

        assembly = {

            "format_version": "1.0",

            "generated_by": "GSPL-01_Rhino_Extractor",

            "generator_version": self.cfg["project"]["version"],

            "project": self.cfg["project"]["name"],

            # Root component (defined later by the engineer)
            "root": None,

            "components": []

        }

        # --------------------------------------------------
        # Process Components
        # --------------------------------------------------

        for component in self.database["components"]:

            xmin = ymin = zmin = float("inf")
            xmax = ymax = zmax = float("-inf")

            valid_bbox = False

            # ----------------------------------------------
            # Compute Component Bounding Box
            # ----------------------------------------------

            for entity in component["entities"]:

                obj = self.model.Objects[entity["index"]]

                bbox = obj.Geometry.GetBoundingBox()

                if not bbox.IsValid:

                    self.logger.warning(
                        "Invalid BoundingBox in component '%s'",
                        component["name"]
                    )

                    continue

                valid_bbox = True

                xmin = min(xmin, bbox.Min.X)
                ymin = min(ymin, bbox.Min.Y)
                zmin = min(zmin, bbox.Min.Z)

                xmax = max(xmax, bbox.Max.X)
                ymax = max(ymax, bbox.Max.Y)
                zmax = max(zmax, bbox.Max.Z)

            # ----------------------------------------------
            # Default values if BoundingBox failed
            # ----------------------------------------------

            if not valid_bbox:

                xmin = ymin = zmin = 0.0
                xmax = ymax = zmax = 0.0

            # ----------------------------------------------
            # Bounding Box Center
            # ----------------------------------------------

            center = [

                round((xmin + xmax) / 2.0, 6),
                round((ymin + ymax) / 2.0, 6),
                round((zmin + zmax) / 2.0, 6)

            ]

            # ----------------------------------------------
            # Bounding Box Information
            # ----------------------------------------------

            bounding_box = {

                "min": [

                    round(xmin, 6),
                    round(ymin, 6),
                    round(zmin, 6)

                ],

                "max": [

                    round(xmax, 6),
                    round(ymax, 6),
                    round(zmax, 6)

                ],

                "center": center,

                "size": [

                    round(xmax - xmin, 6),
                    round(ymax - ymin, 6),
                    round(zmax - zmin, 6)

                ]

            }

            # ----------------------------------------------
            # Default Scene Object (Shape)
            # ----------------------------------------------

            default_shape = {

                "type": "shape",

                "name": component["name"],

                "position": [

                    0.0,
                    0.0,
                    0.0

                ],

                "orientation": [

                    0.0,
                    0.0,
                    0.0

                ],

                "properties": {

                    "dynamic": False,

                    "respondable": False,

                    "visible": True

                }

            }

            # ----------------------------------------------
            # Assembly Component
            # ----------------------------------------------

            assembly["components"].append({

                "id": component["id"],

                "name": component["name"],

                "description": "",

                "notes": "",

                "enabled": True,

                "parent": None,

                "frame": {

                    "position": center,

                    "orientation": [

                        0.0,
                        0.0,
                        0.0

                    ]

                },

                "bounding_box": bounding_box,

                "models": [

                    "visual"

                ],

                "objects": [

                    default_shape

                ]

            })

        # --------------------------------------------------
        # Sort Components by ID
        # --------------------------------------------------

        assembly["components"] = sorted(

            assembly["components"],

            key=lambda c: c["id"]

        )

        # --------------------------------------------------
        # Output File
        # --------------------------------------------------

        output_file = (

            Path(self.cfg["paths"]["database_directory"])

            /

            self.cfg["files"]["assembly"]

        )

        # --------------------------------------------------
        # Save JSON
        # --------------------------------------------------

        try:

            with open(output_file, "w", encoding="utf-8") as f:

                json.dump(

                    assembly,

                    f,

                    indent=4,

                    ensure_ascii=False

                )

            self.logger.info("")
            self.logger.info("Assembly template successfully saved.")
            self.logger.info("File       : %s", output_file)
            self.logger.info("Components : %d", len(assembly["components"]))

        except Exception as e:

            self.logger.error("")
            self.logger.error("Unable to save assembly template.")
            self.logger.error(str(e))
            raise

        self.logger.info("========================================================")

 
    # ------------------------------------------------------
    # Generate Assembly Excel Table
    # ------------------------------------------------------
    def update_parent_ids(self, workbook):

        self.logger.info("Updating Parent IDs...")

        ws = workbook["Components"]

        # --------------------------------------------
        # Build Name -> ID dictionary
        # --------------------------------------------

        component_map = {}

        for row in range(2, ws.max_row + 1):

            component_id = ws[f"A{row}"].value
            component_name = ws[f"B{row}"].value

            component_map[component_name] = component_id

        # --------------------------------------------
        # Update Parent IDs
        # --------------------------------------------

        for row in range(2, ws.max_row + 1):

            parent_name = ws[f"F{row}"].value

            if parent_name is None or str(parent_name).strip() == "":

                ws[f"G{row}"].value = ""

                continue

            if parent_name not in component_map:

                self.logger.warning(
                    "Unknown parent '%s' in row %d",
                    parent_name,
                    row
                )

                ws[f"G{row}"].value = ""

                continue

            ws[f"G{row}"].value = component_map[parent_name]

        self.logger.info("Parent IDs updated.")
    def create_components_sheet(self, workbook):

        self.logger.info("Creating worksheet: Components")

        ws = workbook.create_sheet("Components")

        # --------------------------------------------------
        # Header
        # --------------------------------------------------

        headers = [

            "ID",
            "Name",
            "Type",
            "Description",
            "Enabled",
            "Parent Name",
            "Parent ID",
            "Visual",
            "Simulation",
            "Reviewed",
            "Status",
            "Notes"

        ]

        for col, header in enumerate(headers, start=1):

            ws.cell(row=1, column=col).value = header

        # --------------------------------------------------
        # Components
        # --------------------------------------------------

        row = 2

        for component in self.database["components"]:

            # ID
            ws.cell(row=row, column=1).value = component["id"]

            # Name
            ws.cell(row=row, column=2).value = component["name"]

            # Type (Engineer)
            ws.cell(row=row, column=3).value = ""

            # Description
            ws.cell(row=row, column=4).value = ""

            # Enabled
            ws.cell(row=row, column=5).value = True

            # Parent Name (Engineer)
            ws.cell(row=row, column=6).value = ""

            # Parent ID (Filled automatically by GSPL-02)
            ws.cell(row=row, column=7).value = ""

            # Visual
            ws.cell(row=row, column=8).value = True

            # Simulation
            ws.cell(row=row, column=9).value = False

            # Reviewed
            ws.cell(row=row, column=10).value = False

            # Status
            ws.cell(row=row, column=11).value = "NEW"

            # Notes
            ws.cell(row=row, column=12).value = ""

            row += 1

        # --------------------------------------------------
        # Freeze Header
        # --------------------------------------------------

        ws.freeze_panes = "A2"

        # --------------------------------------------------
        # Auto Filter
        # --------------------------------------------------

        ws.auto_filter.ref = f"A1:L{row-1}"

        # --------------------------------------------------
        # Column Widths
        # --------------------------------------------------

        widths = {

            "A": 8,
            "B": 35,
            "C": 18,
            "D": 35,
            "E": 12,
            "F": 35,
            "G": 12,
            "H": 12,
            "I": 14,
            "J": 12,
            "K": 15,
            "L": 40

        }

        for col, width in widths.items():

            ws.column_dimensions[col].width = width

        self.logger.info(
            "Components worksheet created (%d components).",
            len(self.database["components"])
        )
    def create_objects_sheet(self, workbook):

        self.logger.info("Creating worksheet: Objects")

        ws = workbook.create_sheet("Objects")

        # --------------------------------------------------
        # Header
        # --------------------------------------------------

        headers = [

            "Component ID",
            "Component Name",

            "Object Name",
            "Object Type",

            "Enabled",

            "Pos X",
            "Pos Y",
            "Pos Z",

            "Rot X",
            "Rot Y",
            "Rot Z",

            "Joint Type",
            "Cyclic",
            "Lower Limit",
            "Upper Limit",
            "Motor",

            "Notes"

        ]

        for col, header in enumerate(headers, start=1):

            ws.cell(row=1, column=col).value = header

        # --------------------------------------------------
        # Default Shape
        # --------------------------------------------------

        row = 2

        for component in self.database["components"]:

            ws.cell(row=row, column=1).value = component["id"]

            ws.cell(row=row, column=2).value = component["name"]

            ws.cell(row=row, column=3).value = component["name"]

            ws.cell(row=row, column=4).value = "shape"

            ws.cell(row=row, column=5).value = True

            # Position
            ws.cell(row=row, column=6).value = 0.0
            ws.cell(row=row, column=7).value = 0.0
            ws.cell(row=row, column=8).value = 0.0

            # Orientation
            ws.cell(row=row, column=9).value = 0.0
            ws.cell(row=row, column=10).value = 0.0
            ws.cell(row=row, column=11).value = 0.0

            # Joint Properties
            ws.cell(row=row, column=12).value = ""
            ws.cell(row=row, column=13).value = ""
            ws.cell(row=row, column=14).value = ""
            ws.cell(row=row, column=15).value = ""
            ws.cell(row=row, column=16).value = ""

            # Notes
            ws.cell(row=row, column=17).value = ""

            row += 1

        # --------------------------------------------------
        # Freeze Header
        # --------------------------------------------------

        ws.freeze_panes = "A2"

        # --------------------------------------------------
        # Auto Filter
        # --------------------------------------------------

        ws.auto_filter.ref = f"A1:Q{row-1}"

        # --------------------------------------------------
        # Column Widths
        # --------------------------------------------------

        widths = {

            "A": 15,
            "B": 30,
            "C": 30,
            "D": 18,
            "E": 12,

            "F": 10,
            "G": 10,
            "H": 10,

            "I": 10,
            "J": 10,
            "K": 10,

            "L": 18,
            "M": 10,
            "N": 14,
            "O": 14,
            "P": 10,

            "Q": 40

        }

        for col, width in widths.items():

            ws.column_dimensions[col].width = width

        self.logger.info(
            "Objects worksheet created (%d default shapes).",
            len(self.database["components"])
        )
    def create_lists_sheet(self, workbook):

        self.logger.info("Creating worksheet: Lists")

        ws = workbook.create_sheet("Lists")

        # ==================================================
        # Components
        # ==================================================

        ws["A1"] = "Component ID"
        ws["B1"] = "Component Name"

        row = 2

        for component in self.database["components"]:

            ws.cell(row=row, column=1).value = component["id"]
            ws.cell(row=row, column=2).value = component["name"]

            row += 1

        # ==================================================
        # Boolean
        # ==================================================

        ws["C1"] = "Boolean"

        ws["C2"] = True
        ws["C3"] = False

        # ==================================================
        # Object Types
        # ==================================================

        ws["D1"] = "Object Type"

        object_types = [

            "shape",
            "joint",
            "dummy",
            "camera",
            "light",
            "vision_sensor",
            "proximity_sensor",
            "force_sensor",
            "path",
            "point_cloud",
            "octree"

        ]

        row = 2

        for value in object_types:

            ws.cell(row=row, column=4).value = value
            row += 1

        # ==================================================
        # Joint Types
        # ==================================================

        ws["E1"] = "Joint Type"

        joint_types = [

            "revolute",
            "prismatic",
            "spherical"

        ]

        row = 2

        for value in joint_types:

            ws.cell(row=row, column=5).value = value
            row += 1

        # ==================================================
        # Project Status
        # ==================================================

        ws["F1"] = "Status"

        status = [

            "NEW",
            "IN_PROGRESS",
            "REVIEWED",
            "VERIFIED",
            "DONE"

        ]

        row = 2

        for value in status:

            ws.cell(row=row, column=6).value = value
            row += 1

        # ==================================================
        # Component Types
        # ==================================================

        ws["G1"] = "Component Type"

        component_types = [

            "Mechanical",
            "Electrical",
            "Electronic",
            "Sensor",
            "Actuator",
            "Structure",
            "Fastener",
            "Other"

        ]

        row = 2

        for value in component_types:

            ws.cell(row=row, column=7).value = value
            row += 1

        # ==================================================
        # Motor Mode
        # ==================================================

        ws["H1"] = "Motor Mode"

        motor_modes = [

            "Position",
            "Velocity",
            "Torque"

        ]

        row = 2

        for value in motor_modes:

            ws.cell(row=row, column=8).value = value
            row += 1

        # ==================================================
        # Sensor Types
        # ==================================================

        ws["I1"] = "Sensor Type"

        sensor_types = [

            "Vision",
            "Proximity",
            "Force",
            "Hall",
            "Lidar",
            "IMU",
            "GPS",
            "Other"

        ]

        row = 2

        for value in sensor_types:

            ws.cell(row=row, column=9).value = value
            row += 1

        # ==================================================
        # Light Types
        # ==================================================

        ws["J1"] = "Light Type"

        light_types = [

            "Omnidirectional",
            "Spot",
            "Directional"

        ]

        row = 2

        for value in light_types:

            ws.cell(row=row, column=10).value = value
            row += 1

        # ==================================================
        # Dynamic
        # ==================================================

        ws["K1"] = "Dynamic"

        ws["K2"] = True
        ws["K3"] = False

        # ==================================================
        # Respondable
        # ==================================================

        ws["L1"] = "Respondable"

        ws["L2"] = True
        ws["L3"] = False

        # ==================================================
        # Visible
        # ==================================================

        ws["M1"] = "Visible"

        ws["M2"] = True
        ws["M3"] = False

        # ==================================================
        # Control Loop
        # ==================================================

        ws["N1"] = "Control Loop"

        ws["N2"] = True
        ws["N3"] = False

        # ==================================================
        # Cyclic
        # ==================================================

        ws["O1"] = "Cyclic"

        ws["O2"] = True
        ws["O3"] = False

        # ==================================================
        # Hide Worksheet
        # ==================================================

        ws.sheet_state = "veryHidden"

        self.logger.info("Lists worksheet created.")
    def apply_excel_styles(self, workbook):

        self.logger.info("Applying Excel styles")

        # --------------------------------------------------
        # Colors
        # --------------------------------------------------

        HEADER_FILL = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        AUTO_FILL = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE"
        )

        EDIT_FILL = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

        CALC_FILL = PatternFill(
            fill_type="solid",
            fgColor="E7E6E6"
        )

        HEADER_FONT = Font(
            bold=True,
            color="FFFFFF"
        )

        THIN = Side(
            border_style="thin",
            color="A6A6A6"
        )

        BORDER = Border(
            left=THIN,
            right=THIN,
            top=THIN,
            bottom=THIN
        )

        CENTER = Alignment(
            horizontal="center",
            vertical="center"
        )

        LEFT = Alignment(
            horizontal="left",
            vertical="center"
        )

        # ==================================================
        # Components
        # ==================================================

        ws = workbook["Components"]

        # Header

        for cell in ws[1]:

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Automatic Columns

        for col in ["A", "B", "G"]:

            for cell in ws[col][1:]:

                cell.fill = AUTO_FILL
                cell.border = BORDER

        # Editable Columns

        for col in ["C","D","E","F","H","I","J","K","L"]:

            for cell in ws[col][1:]:

                cell.fill = EDIT_FILL
                cell.border = BORDER

        # ==================================================
        # Objects
        # ==================================================

        ws = workbook["Objects"]

        for cell in ws[1]:

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Automatic

        for col in ["A","B"]:

            for cell in ws[col][1:]:

                cell.fill = AUTO_FILL
                cell.border = BORDER

        # Editable

        for col in [

            "C","D","E",

            "F","G","H",

            "I","J","K",

            "L","M","N","O","P",

            "Q"

        ]:

            for cell in ws[col][1:]:

                cell.fill = EDIT_FILL
                cell.border = BORDER

        self.logger.info("Excel styles applied.")
    def apply_excel_validations(self, workbook):

        self.logger.info("Applying Excel validations")

        # ==================================================
        # Worksheets
        # ==================================================

        ws_components = workbook["Components"]
        ws_objects = workbook["Objects"]

        component_count = len(self.database["components"])

        # ==================================================
        # Data Validations
        # ==================================================

        dv_boolean = DataValidation(
            type="list",
            formula1="=Lists!$C$2:$C$3",
            allow_blank=False
        )

        dv_parent = DataValidation(
            type="list",
            formula1=f"=Lists!$B$2:$B${component_count+1}",
            allow_blank=True
        )

        dv_component_type = DataValidation(
            type="list",
            formula1="=Lists!$G$2:$G$100",
            allow_blank=True
        )

        dv_status = DataValidation(
            type="list",
            formula1="=Lists!$F$2:$F$100",
            allow_blank=True
        )

        dv_object_type = DataValidation(
            type="list",
            formula1="=Lists!$D$2:$D$100",
            allow_blank=False
        )

        dv_joint_type = DataValidation(
            type="list",
            formula1="=Lists!$E$2:$E$100",
            allow_blank=True
        )

        dv_motor_mode = DataValidation(
            type="list",
            formula1="=Lists!$H$2:$H$100",
            allow_blank=True
        )

        dv_sensor_type = DataValidation(
            type="list",
            formula1="=Lists!$I$2:$I$100",
            allow_blank=True
        )

        dv_light_type = DataValidation(
            type="list",
            formula1="=Lists!$J$2:$J$100",
            allow_blank=True
        )

        # ==================================================
        # Register Validations
        # ==================================================

        validations = [

            dv_boolean,
            dv_parent,
            dv_component_type,
            dv_status,

            dv_object_type,
            dv_joint_type,
            dv_motor_mode,
            dv_sensor_type,
            dv_light_type

        ]

        for dv in validations:

            ws_components.add_data_validation(dv)
            ws_objects.add_data_validation(dv)

        # ==================================================
        # Components Sheet
        # ==================================================

        for row in range(2, ws_components.max_row + 1):

            # Component Type
            dv_component_type.add(f"C{row}")

            # Enabled
            dv_boolean.add(f"E{row}")

            # Parent Name
            dv_parent.add(f"F{row}")

            # Visual
            dv_boolean.add(f"H{row}")

            # Simulation
            dv_boolean.add(f"I{row}")

            # Reviewed
            dv_boolean.add(f"J{row}")

            # Status
            dv_status.add(f"K{row}")

        # ==================================================
        # Objects Sheet
        # ==================================================

        for row in range(2, ws_objects.max_row + 1):

            # Object Type
            dv_object_type.add(f"D{row}")

            # Enabled
            dv_boolean.add(f"E{row}")

            # Joint Type
            dv_joint_type.add(f"L{row}")

            # Cyclic
            dv_boolean.add(f"M{row}")

            # Motor
            dv_boolean.add(f"P{row}")

            # --------------------------------------------------
            # Reserved for future versions
            # --------------------------------------------------
            #
            # When new columns are added:
            #
            # Motor Mode  -> dv_motor_mode
            # Sensor Type -> dv_sensor_type
            # Light Type  -> dv_light_type
            #
            # The semantic validation (e.g. requiring Joint Type
            # only when Object Type == "joint") will be performed
            # by GSPL-02_Assembly_Builder.py.
            #
            # --------------------------------------------------

        self.logger.info("Excel validations successfully applied.")
    def save_assembly_table(self, workbook):

        self.logger.info("Saving Assembly Table")

        # --------------------------------------------------
        # Workbook Properties
        # --------------------------------------------------

        workbook.properties.creator = "GSPL-01_Rhino_Extractor"

        workbook.properties.title = "GSPL Assembly Table"

        workbook.properties.subject = "GIAR Simulation Pipeline for LiDAR"

        workbook.properties.description = (
            "Assembly definition table automatically generated "
            "by GSPL-01_Rhino_Extractor."
        )

        workbook.properties.company = "GIAR"

        workbook.properties.category = "Engineering"

        workbook.properties.keywords = (
            "GSPL, Rhino, CoppeliaSim, Assembly, LiDAR"
        )

        # --------------------------------------------------
        # Output File
        # --------------------------------------------------

        output_file = (

            Path(self.cfg["paths"]["database_directory"])

            /

            self.cfg["files"]["assembly_table"]

        )

        # --------------------------------------------------
        # Save Workbook
        # --------------------------------------------------

        try:

            workbook.save(output_file)

            self.logger.info("")
            self.logger.info("Assembly table successfully saved.")
            self.logger.info("File : %s", output_file)

        except Exception as e:

            self.logger.error("")
            self.logger.error("Unable to save assembly table.")
            self.logger.error(str(e))
            raise

    def generate_assembly_table(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Generating Assembly Table")
        self.logger.info("========================================================")

        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        self.create_components_sheet(workbook)

        self.create_objects_sheet(workbook)

        self.create_lists_sheet(workbook)

        self.apply_excel_styles(workbook)

        self.apply_excel_validations(workbook)

        self.save_assembly_table(workbook)

        self.logger.info("Assembly table successfully generated.")
        self.logger.info("========================================================")

    # ------------------------------------------------------
    # Run
    # ------------------------------------------------------
    def run(self):

        self.load_config()
        self.configure_logger()
        self.print_banner()
        self.verify_project()
        self.open_model()
        self.audit_geometry()
        self.print_model_information()
        self.inspect_objects()
        self.build_database()
        self.save_database()
        self.generate_assembly_template()
        self.generate_assembly_table()
 
        self.logger.info("")
        self.logger.info("Finished.")



# ==========================================================
# Main
# ==========================================================

def main():

    extractor = RhinoExtractor()

    extractor.run()


if __name__ == "__main__":

    main()