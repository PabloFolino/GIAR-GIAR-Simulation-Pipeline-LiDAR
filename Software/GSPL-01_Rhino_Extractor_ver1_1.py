#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
------------------------------------------------------------
GSPL-01_Rhino_Extractor.py

GIAR Simulation Pipeline for LiDAR

Version : 1.1

Author:
    Ing. Pablo Daniel Folino

Description:
    Opens a Rhino (.3dm) model and inspects its contents.

------------------------------------------------------------
"""

from pathlib import Path
import json
import logging
import sys
import rhino3dm
import uuid
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime
import copy


CONFIG_FILE = "config.json"


# ==========================================================
# Rhino Extractor
# ==========================================================

class RhinoExtractor:

    def __init__(self):

        self.cfg = None
        self.logger = None
        self.model = None


    # ------------------------------------------------------
    # Configuration
    # ------------------------------------------------------
    def load_config(self):

        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            self.cfg = json.load(f)


    # ------------------------------------------------------
    # Logger
    # ------------------------------------------------------
    def configure_logger(self):

        log_dir = Path(self.cfg["paths"]["log_directory"])

        log_dir.mkdir(parents=True, exist_ok=True)

        logfile = log_dir / "GSPL-01.log"

        logging.basicConfig(

            level=logging.INFO,

            format="%(asctime)s | %(levelname)-8s | %(message)s",

            handlers=[
                logging.FileHandler(logfile, mode="w"),
                logging.StreamHandler(sys.stdout)
            ]
        )

        self.logger = logging.getLogger("GSPL")

    # ------------------------------------------------------
    # Banner
    # ------------------------------------------------------
    def print_banner(self):

        self.logger.info("")
        self.logger.info("====================================================")
        self.logger.info(" GSPL-01 Rhino Extractor")
        self.logger.info(" Version : 1.1")
        self.logger.info("====================================================")
        self.logger.info(" Author:")
        self.logger.info("     Ing. Pablo Daniel Folino")
        self.logger.info("====================================================")
        self.logger.info("")

        self.logger.info("Project : %s",
                         self.cfg["project"]["name"])

        self.logger.info("Version : %s",
                         self.cfg["project"]["version"])

        self.logger.info("")


    # ------------------------------------------------------
    # Verify project
    # ------------------------------------------------------
    def verify_project(self):

        for directory in [

            self.cfg["paths"]["input_directory"],
            self.cfg["paths"]["stl_directory"],
            self.cfg["paths"]["database_directory"],
            self.cfg["paths"]["output_directory"]

        ]:

            if Path(directory).exists():

                self.logger.info("[ OK ] %s", directory)

            else:

                self.logger.error("[ERROR] %s", directory)

    # ------------------------------------------------------
    # Open Rhino Model
    # ------------------------------------------------------
    def open_model(self):

        filename = (

            Path(self.cfg["paths"]["input_directory"])

            /

            self.cfg["files"]["input_file_3dm"]

        )

        self.logger.info("")
        self.logger.info("Opening Rhino model...")
        self.logger.info("%s", filename)

        self.model = rhino3dm.File3dm.Read(str(filename))

        if self.model is None:

            raise RuntimeError("Unable to open Rhino file.")

        self.logger.info("[ OK ] Rhino model loaded.")

    # ------------------------------------------------------
    # Model information
    # ------------------------------------------------------
    def print_model_information(self):

        self.logger.info("")
        self.logger.info("----------------------------------------")
        self.logger.info("Model Information")
        self.logger.info("----------------------------------------")

        self.logger.info("File Version : %s",
                         self.model.Settings.ModelUnitSystem)

        self.logger.info("Layers       : %d",
                         len(self.model.Layers))

        self.logger.info("Objects      : %d",
                         len(self.model.Objects))

        self.logger.info("Groups       : %d",
                         len(self.model.Groups))

    # ------------------------------------------------------
    # Inspect Objects
    # ------------------------------------------------------
    def inspect_objects(self):

        self.logger.info("")
        self.logger.info("----------------------------------------")
        self.logger.info("Objects")
        self.logger.info("----------------------------------------")

        for i, obj in enumerate(self.model.Objects, start=1):

            att = obj.Attributes

            geo = obj.Geometry

            self.logger.info("")
            self.logger.info("[%04d]", i)

            self.logger.info("Name     : %s", att.Name)

            self.logger.info("GUID     : %s", att.Id)

            self.logger.info("Layer Id : %d", att.LayerIndex)

            self.logger.info("Geometry : %s",
                             geo.ObjectType)

    # ------------------------------------------------------
    # Geometry Audit
    # ------------------------------------------------------
    def audit_geometry(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" GEOMETRY AUDIT")
        self.logger.info("========================================================")

        errors = 0

        for obj in self.model.Objects:

            att = obj.Attributes
            geo = obj.Geometry

            name = att.Name if att.Name else "<UNNAMED>"

            # Obtener nombre de la Layer
            try:
                layer = self.model.Layers[att.LayerIndex].Name
            except:
                layer = "<UNKNOWN>"

            guid = att.Id

            # --------------------------------------------------
            # BREP
            # --------------------------------------------------

            if isinstance(geo, rhino3dm.Brep):

                if geo.IsSolid:

                    continue

                errors += 1

                self.logger.error("")
                self.logger.error("ERROR %03d", errors)
                self.logger.error("--------------------------------------------")
                self.logger.error("Component : %s", name)
                self.logger.error("Layer     : %s", layer)
                self.logger.error("GUID      : %s", guid)
                self.logger.error("Geometry  : Open Brep")
                self.logger.error("")
                self.logger.error("Action:")
                self.logger.error("    Convert this Brep into a CLOSED solid.")
                self.logger.error("--------------------------------------------")

                continue

            # --------------------------------------------------
            # EXTRUSION
            # --------------------------------------------------

            if isinstance(geo, rhino3dm.Extrusion):

                if geo.IsSolid:

                    continue

                errors += 1

                self.logger.error("")
                self.logger.error("ERROR %03d", errors)
                self.logger.error("--------------------------------------------")
                self.logger.error("Component : %s", name)
                self.logger.error("Layer     : %s", layer)
                self.logger.error("GUID      : %s", guid)
                self.logger.error("Geometry  : Open Extrusion")
                self.logger.error("")
                self.logger.error("Action:")
                self.logger.error("    Close the extrusion.")
                self.logger.error("--------------------------------------------")

                continue

            # --------------------------------------------------
            # INVALID GEOMETRY
            # --------------------------------------------------

            errors += 1

            self.logger.error("")
            self.logger.error("ERROR %03d", errors)
            self.logger.error("--------------------------------------------")
            self.logger.error("Component : %s", name)
            self.logger.error("Layer     : %s", layer)
            self.logger.error("GUID      : %s", guid)
            self.logger.error("Geometry  : %s", geo.ObjectType)
            self.logger.error("")
            self.logger.error("Action:")
            self.logger.error("    Delete or convert this object into a closed solid.")
            self.logger.error("--------------------------------------------")

        self.logger.info("")
        self.logger.info("========================================================")

        if errors == 0:

            self.logger.info("Geometry Audit : PASSED")

        else:

            self.logger.error("Geometry Audit : FAILED")
            self.logger.error("Total Errors   : %d", errors)

            raise RuntimeError(
                f"Geometry Audit FAILED ({errors} errors)"
            )

        self.logger.info("========================================================")

    # ------------------------------------------------------
    # Initialize Database
    # ------------------------------------------------------
    def initialize_database(self):

        self.database = {

            "project": {

                "name": self.cfg["project"]["name"],
                "version": self.cfg["project"]["version"]

            },

            "statistics": {

                "components": 0,
                "entities": 0

            },

            "components": []

        }
    
    # ------------------------------------------------------
    # Build Component Database
    # ------------------------------------------------------
    def build_database(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Building Component Database")
        self.logger.info("========================================================")

        self.initialize_database()

        components = {}

        for index, obj in enumerate(self.model.Objects):

            att = obj.Attributes
            geo = obj.Geometry

            # --------------------------------------------------
            # Component Name
            # --------------------------------------------------
            name = att.Name.strip() if att.Name else "<UNNAMED>"

            # --------------------------------------------------
            # Layer
            # --------------------------------------------------
            try:
                layer = self.model.Layers[att.LayerIndex].Name
            except Exception:
                layer = "<UNKNOWN>"

            # --------------------------------------------------
            # Geometry Type
            # --------------------------------------------------
            geometry = str(geo.ObjectType).replace("ObjectType.", "")

            # --------------------------------------------------
            # Create Component
            # --------------------------------------------------

            if name not in components:

                component = {
                    "component_id": len(components) + 1,
                    "uuid": str(uuid.uuid4()),
                    "name": name,
                    "type":"Mechanical",
                    "layer": layer,
                    "parent": None,
                    "children": [],
                    "stl": None,
                    "frame": None,
                    "bounding_box": None,
                    "mass": None,
                    "material": None,
                    "properties": {
                        "dynamic": False,
                        "respondable": False,
                        "visible": True,
                        "collidable": False,
                        "detectable": False,
                        "measurable": False
                    },
                    "status": {
                        "geometry_audit": "PASSED",
                        "hierarchy_audit": "PENDING",
                        "stl_export": "PENDING",
                        "simulation": "PENDING"
                    },
                    "entities": []
                }

                components[name] = component

            # --------------------------------------------------
            # Add Entity
            # --------------------------------------------------
            entity = {
                "index": index,
                "guid": str(att.Id),
                "geometry": geometry
            }
            components[name]["entities"].append(entity)

        # ------------------------------------------------------
        # Sort Components
        # ------------------------------------------------------
        component_list = sorted(
            components.values(),
            key=lambda c: c["name"].lower()
        )

        # ------------------------------------------------------
        # Reassign IDs after sorting
        # ------------------------------------------------------
        for new_id, component in enumerate(component_list, start=1):
            component["component_id"] = new_id

        # ------------------------------------------------------
        # Statistics
        # ------------------------------------------------------
        self.database["statistics"]["components"] = len(component_list)
        self.database["statistics"]["entities"] = len(self.model.Objects)
        self.database["components"] = component_list

        # ------------------------------------------------------
        # Log
        # ------------------------------------------------------
        self.logger.info("Components : %d", len(component_list))
        self.logger.info("Entities   : %d", len(self.model.Objects))

        self.logger.info("")
        self.logger.info("Component List")
        self.logger.info("--------------------------------------------------------")

        for component in component_list:

            self.logger.info(
                "[%03d] %-35s (%3d entities)",
                component["component_id"],
                component["name"],
                len(component["entities"])
            )

        self.logger.info("--------------------------------------------------------")

    # ------------------------------------------------------
    # Build Geometry Database
    # ------------------------------------------------------
    def build_geometry_database(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Building Geometry Database")
        self.logger.info("========================================================")

        for component in self.database["components"]:

            xmin = ymin = zmin = float("inf")
            xmax = ymax = zmax = float("-inf")

            valid_bbox = False

            # --------------------------------------------------
            # Compute Bounding Box
            # --------------------------------------------------

            for entity in component["entities"]:

                obj = self.model.Objects[entity["index"]]

                bbox = obj.Geometry.GetBoundingBox()

                if not bbox.IsValid:

                    self.logger.warning(
                        "Invalid BoundingBox in component '%s'",
                        component["name"]
                    )

                    continue

                valid_bbox = True

                xmin = min(xmin, bbox.Min.X)
                ymin = min(ymin, bbox.Min.Y)
                zmin = min(zmin, bbox.Min.Z)

                xmax = max(xmax, bbox.Max.X)
                ymax = max(ymax, bbox.Max.Y)
                zmax = max(zmax, bbox.Max.Z)

            # --------------------------------------------------
            # Default Bounding Box
            # --------------------------------------------------

            if not valid_bbox:

                xmin = ymin = zmin = 0.0
                xmax = ymax = zmax = 0.0

            # --------------------------------------------------
            # Center
            # --------------------------------------------------

            center = [

                round((xmin + xmax) / 2.0, 6),
                round((ymin + ymax) / 2.0, 6),
                round((zmin + zmax) / 2.0, 6)

            ]

            # --------------------------------------------------
            # Bounding Box
            # --------------------------------------------------

            component["bounding_box"] = {

                "min": [
                    round(xmin, 6),
                    round(ymin, 6),
                    round(zmin, 6)
                ],

                "max": [
                    round(xmax, 6),
                    round(ymax, 6),
                    round(zmax, 6)
                ],

                "center": center,

                "size": [
                    round(xmax - xmin, 6),
                    round(ymax - ymin, 6),
                    round(zmax - zmin, 6)
                ]

            }

            # --------------------------------------------------
            # Component Frame
            # --------------------------------------------------

            component["frame"] = {
                "name": "Component Frame",
                "in_respect_of": "parent",
                "position": center,
                "orientation": [
                    0.0,
                    0.0,
                    0.0
                ]
            }

        self.logger.info(
            "Geometry computed for %d components.",
            len(self.database["components"])
        )

    # ------------------------------------------------------
    # Save Database
    # ------------------------------------------------------
    def save_database(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Saving Component Database")
        self.logger.info("========================================================")

        # --------------------------------------------------
        # Output directory
        # --------------------------------------------------

        database_dir = Path(self.cfg["paths"]["database_directory"])

        database_dir.mkdir(parents=True, exist_ok=True)

        # --------------------------------------------------
        # Output file
        # --------------------------------------------------

        output_file = (
            database_dir /
            self.cfg["files"]["model_database"]
        )

        # --------------------------------------------------
        # Save JSON
        # --------------------------------------------------

        try:

            with open(output_file, "w", encoding="utf-8") as f:

                json.dump(
                    self.database,
                    f,
                    indent=4,
                    ensure_ascii=False
                )

            self.logger.info("")
            self.logger.info("[ OK ] Database successfully saved.")
            self.logger.info("File       : %s", output_file)
            self.logger.info("Components : %d",
                            self.database["statistics"]["components"])
            self.logger.info("Entities   : %d",
                            self.database["statistics"]["entities"])

        except Exception as e:

            self.logger.error("")
            self.logger.error("[ERROR] Unable to save database.")
            self.logger.error(str(e))
            raise

        self.logger.info("========================================================")

    # ------------------------------------------------------
    # Generate Assembly Template
    # ------------------------------------------------------
    def generate_assembly_template(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Generating Assembly Template")
        self.logger.info("========================================================")

        now = datetime.now()

        # --------------------------------------------------
        # Assembly Header
        # --------------------------------------------------
        assembly = {
            "format_version": "1.1",
            "generated_by": "GSPL-01_Rhino_Extractor",
            "generator_version": self.cfg["project"]["version"],
            "generated": {
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%H:%M:%S")
            },
            "project": self.cfg["project"]["name"],
            "root_component_id": None,
            "statistics": {},
            "components": []
        }

        # --------------------------------------------------
        # Root Component
        # --------------------------------------------------
        root = None

        # --------------------------------------------------
         # Object ID Generator
         # --------------------------------------------------
        next_object_id = 1

        # --------------------------------------------------
        # Components
        # --------------------------------------------------
        for source in self.database["components"]:
            if root is None:
                root = source["component_id"]
            # ----------------------------------------------
            # Assign the object number.
            # ----------------------------------------------
            object_id = next_object_id
            # ----------------------------------------------
            # Object counter increment
            # ----------------------------------------------
            next_object_id += 1
            # ----------------------------------------------
            # Default Shape
            # ----------------------------------------------
            default_shape = {
                # Unique identifier of the simulation object.
                # One component may contain one or more simulation objects.
                "object_id": object_id,   
                "component_id": source["component_id"],
                "type": "Shape",
                "name": f'{source["name"]}_Shape',
                "position": [0.0,0.0,0.0],
                "orientation": [0.0,0.0,0.0],
                "properties": {
                    "dynamic": False,
                    "respondable": False,
                    "visible": True,
                    "collidable": False,
                    "detectable": False,
                    "measurable": False
                }
            }
  
            # ----------------------------------------------
            # Assembly Component
            # ----------------------------------------------
            component = {
                "component_id": source["component_id"],
                "name": source["name"],
                "description": "",
                "notes": "",
                "enabled": True,
                "parent": None,
                "frame": copy.deepcopy(
                    source["frame"]
                ),
                "bounding_box": copy.deepcopy(
                    source["bounding_box"]
                ),
                "models": {
                    "visual":True,
                    "simulation":False
                },
                "objects": [
                    default_shape
                ]
            }

            assembly["components"].append(component)

        # --------------------------------------------------
        # Root
        # --------------------------------------------------
        assembly["root_component_id"] = root

        # --------------------------------------------------
        # Statistics
        # --------------------------------------------------
        statistics = {
            "components": len(assembly["components"]),
            "objects": 0,
            "visual_models": 0,
            "simulation_models": 0
        }
        for component in assembly["components"]:
            statistics["objects"] += len(component["objects"])
            if component["models"]["visual"]:
                statistics["visual_models"] += 1
            if component["models"]["simulation"]:
                statistics["simulation_models"] += 1
        assembly["statistics"] = statistics

        # --------------------------------------------------
        # Output File
        # --------------------------------------------------
        output_file = (
            Path(self.cfg["paths"]["database_directory"])
            /
            self.cfg["files"]["assembly"]
        )

        # --------------------------------------------------
        # Save JSON
        # --------------------------------------------------
        with open(
            output_file,
            "w",
            encoding="utf-8"
        ) as f:
            json.dump(
                assembly,
                f,
                indent=4,
                ensure_ascii=False
            )

        # --------------------------------------------------
        # Log
        # --------------------------------------------------
        self.logger.info("Assembly Components : %d", statistics["components"])
        self.logger.info("Objects             : %d", statistics["objects"])
        self.logger.info("Output File         : %s", output_file)
        self.logger.info("========================================================")
 
    # ------------------------------------------------------
    # Generate Assembly Excel Table
    # ------------------------------------------------------
    def protect_worksheet(self, worksheet, editable_columns):

        self.logger.info(
            "Protecting worksheet: %s",
            worksheet.title
        )

        editable = set(editable_columns)

        # ----------------------------------------------
        # Lock / Unlock Cells
        # ----------------------------------------------

        for row in worksheet.iter_rows():

            for cell in row:

                column = cell.column_letter

                if column in editable:

                    cell.protection = Protection(
                        locked=False
                    )

                else:

                    cell.protection = Protection(
                        locked=True
                    )

        # ----------------------------------------------
        # Enable Sheet Protection
        # ----------------------------------------------

        #worksheet.protection.sheet = True
        #worksheet.protection.password = "UTN"
        #worksheet.protection.insertRows = True
       
        #worksheet.protection.enable()

        self.logger.info(
            "Worksheet '%s' prepared.",
            worksheet.title
        )
        
    def update_parent_ids(self, workbook):

        self.logger.info("Updating Parent IDs...")

        ws = workbook["Components"]

        # --------------------------------------------
        # Build Name -> ID dictionary
        # --------------------------------------------

        component_map = {}

        for row in range(2, ws.max_row + 1):

            component_id = ws[f"A{row}"].value
            component_name = ws[f"B{row}"].value

            component_map[str(component_name).strip()] = component_id

        # --------------------------------------------
        # Update Parent IDs
        # --------------------------------------------

        for row in range(2, ws.max_row + 1):

            parent_name = ws[f"F{row}"].value

            if parent_name is None or str(parent_name).strip() == "":

                ws[f"G{row}"].value = ""

                continue

            parent_name = str(parent_name).strip()

            if parent_name not in component_map:

                self.logger.warning(
                    "Unknown parent '%s' in row %d",
                    parent_name,
                    row
                )

                ws[f"G{row}"].value = ""

                continue

            ws[f"G{row}"].value = component_map[parent_name]

        self.logger.info("Parent IDs updated.")
    def create_components_sheet(self, workbook):

        self.logger.info("Creating worksheet: Components")
        ws = workbook.create_sheet("Components")

        # --------------------------------------------------
        # Header
        # --------------------------------------------------
        headers = [
            "Component ID",
            "Name",
            "Type",
            "Description",
            "Enabled",
            "Parent Name",
            "Parent ID",
            "Visual",
            "Simulation",
            "Status",
            "Notes"
        ]

        for col, header in enumerate(headers, start=1):

            ws.cell(row=1, column=col).value = header

        # --------------------------------------------------
        # Components
        # --------------------------------------------------

        row = 2

        for component in self.database["components"]:

            # ID
            ws.cell(row=row, column=1).value = component["component_id"]

            # Name
            ws.cell(row=row, column=2).value = component["name"]

            # Type (Engineer)
            ws.cell(row=row, column=3).value = ""

            # Description
            ws.cell(row=row, column=4).value = ""

            # Enabled
            ws.cell(
                row=row,
                column=5
            ).value = f"=OR(H{row},I{row})"

            # Parent Name (Engineer)
            ws.cell(row=row, column=6).value = ""

            # Parent ID (Automatically generated)
            ws.cell(row=row, column=7).value = ""

            # Visual
            ws.cell(row=row, column=8).value = True

            # Simulation
            ws.cell(row=row, column=9).value = False

            # Status
            ws.cell(row=row, column=10).value = "NEW"

            # Notes
            ws.cell(row=row, column=11).value = ""

            row += 1

        # --------------------------------------------------
        # Freeze Header
        # --------------------------------------------------

        ws.freeze_panes = "A2"

        # --------------------------------------------------
        # Auto Filter
        # --------------------------------------------------

        ws.auto_filter.ref = f"A1:K{row-1}"

        # --------------------------------------------------
        # Column Widths
        # --------------------------------------------------

        widths = {

            "A": 8,
            "B": 30,
            "C": 18,
            "D": 35,
            "E": 12,
            "F": 30,
            "G": 12,
            "H": 12,
            "I": 14,
            "J": 15,
            "K": 40
        }

        for col, width in widths.items():

            ws.column_dimensions[col].width = width

        # --------------------------------------------------
        # Protect Worksheet
        # --------------------------------------------------
        self.protect_worksheet( 
            ws,
            editable_columns=[

                "C",   # Type
                "D",   # Description
                "F",   # Parent Name
                "H",   # Visual
                "I",   # Simulation
                "J",   # Status
                "K"    # Notes
            ]
        )


        self.logger.info(
            "Components worksheet created (%d components).",
            len(self.database["components"])
        )
    def create_objects_sheet(self, workbook):

        """
        ----------------------------------------------------------
        Objects Worksheet

        One row represents one CoppeliaSim object belonging to a
        mechanical component.

        Editable fields
            -Object ID
            - Object Name
            - Object Type
            - Enabled
            - Reference Frame
            - Shape Properties
            - Position / Orientation
            - Joint Parameters
            - Notes

        Protected fields
            - Component ID
            - Component Name

        Automatically generated
            - Bounding Box
        ----------------------------------------------------------
        """

        self.logger.info("Creating worksheet: Objects")

        ws = workbook.create_sheet("Objects")

        # --------------------------------------------------
        # Header
        # --------------------------------------------------

        headers = [

            "Component ID",
            "Component Name",

            "Object ID",
            "Object Name",
            "Object Type",

            "Enabled",

            "Reference Frame",

            "Dynamic",
            "Respondable",
            "Visible",
            "Collidable",
            "Detectable",
            "Measurable",

            "Pos X",
            "Pos Y",
            "Pos Z",

            "Rot X",
            "Rot Y",
            "Rot Z",

            "BBox Min X",
            "BBox Min Y",
            "BBox Min Z",

            "BBox Size X",
            "BBox Size Y",
            "BBox Size Z",

            "Joint Type",
            "Cyclic",
            "Lower Limit",
            "Upper Limit",
            "Motor",

            "Notes"

        ]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header

        # --------------------------------------------------
        # Default Shape
        # --------------------------------------------------
        row = 2

        for component in self.database["components"]:
            # ----------------------------------------------
            # Component Frame
            # ----------------------------------------------
            frame = component.get("frame", {})

            reference_frame = frame.get(
                "in_respect_of",
                "parent"
            )

            position = frame.get(
                "position",
                [0.0, 0.0, 0.0]
            )

            orientation = frame.get(
                "orientation",
                [0.0, 0.0, 0.0]
            )

            # ----------------------------------------------
            # Bounding Box
            # ----------------------------------------------
            bbox = component.get("bounding_box", {})
            bbox_min = bbox.get(
                "min",
                [0.0, 0.0, 0.0]
            )
            bbox_size = bbox.get(
                "size",
                [0.0, 0.0, 0.0]
            )

            # ----------------------------------------------
            # Default Properties
            # ----------------------------------------------
            properties = component.get("properties", {})
            dynamic = properties.get(
                "dynamic",
                False
            )
            respondable = properties.get(
                "respondable",
                False
            )
            visible = properties.get(
                "visible",
                True
            )
            collidable = properties.get(
                "collidable",
                False
            )
            detectable = properties.get(
                "detectable",
                False
            )
            measurable = properties.get(
                "measurable",
                False
            )

            # ----------------------------------------------
            # Identification
            # ----------------------------------------------

            ws.cell(row=row, column=1).value = component["component_id"]
            ws.cell(row=row, column=2).value = component["name"]

            # ----------------------------------------------
            # Default Object
            # ----------------------------------------------
            object_type = "Shape"
            ws.cell(
                row=row,
                column=3
            ).value = component["component_id"]
            # Object Name (calculated by Excel)
            ws.cell(
                row=row,
                column=4
            ).value = f"=B{row}&\"_\"&E{row}"
            # Object Type
            ws.cell(
                row=row,
                column=5
            ).value = object_type
            ws.cell(
                row=row,
                column=6
            ).value = f"=OR(H{row},I{row},J{row},K{row},L{row},M{row})"
            ws.cell(row=row, column=7).value = reference_frame

            # ----------------------------------------------
            # Object Properties
            # ----------------------------------------------
            ws.cell(row=row, column=8).value = dynamic
            ws.cell(row=row, column=9).value = respondable
            ws.cell(row=row, column=10).value = visible
            ws.cell(row=row, column=11).value = collidable
            ws.cell(row=row, column=12).value = detectable
            ws.cell(row=row, column=13).value = measurable

            # ----------------------------------------------
            # Position
            # ----------------------------------------------

            ws.cell(row=row, column=14).value = position[0]
            ws.cell(row=row, column=15).value = position[1]
            ws.cell(row=row, column=16).value = position[2]

            # ----------------------------------------------
            # Orientation
            # ----------------------------------------------

            ws.cell(row=row, column=17).value = orientation[0]
            ws.cell(row=row, column=18).value = orientation[1]
            ws.cell(row=row, column=19).value = orientation[2]

            # ----------------------------------------------
            # Bounding Box
            # ----------------------------------------------

            ws.cell(row=row, column=20).value = bbox_min[0]
            ws.cell(row=row, column=21).value = bbox_min[1]
            ws.cell(row=row, column=22).value = bbox_min[2]
            ws.cell(row=row, column=23).value = bbox_size[0]
            ws.cell(row=row, column=24).value = bbox_size[1]
            ws.cell(row=row, column=25).value = bbox_size[2]

            # ----------------------------------------------
            # Joint Parameters
            # ----------------------------------------------

            ws.cell(row=row, column=26).value = ""
            ws.cell(row=row, column=27).value = ""
            ws.cell(row=row, column=28).value = ""
            ws.cell(row=row, column=29).value = ""
            ws.cell(row=row, column=30).value = ""

            # ----------------------------------------------
            # Notes
            # ----------------------------------------------

            ws.cell(row=row, column=31).value = ""

            row += 1

        # --------------------------------------------------
        # Freeze Header
        # --------------------------------------------------

        ws.freeze_panes = "A2"

        # --------------------------------------------------
        # Auto Filter
        # --------------------------------------------------

        ws.auto_filter.ref = f"A1:AE{row-1}"
        # --------------------------------------------------
        # Column Widths
        # --------------------------------------------------
        widths = {

            # Identification
            "A": 15,   # Component ID
            "B": 30,   # Component Name
            "C": 12,   # Object ID

            # Object
            "D": 30,   # Object Name
            "E": 18,   # Object Type
            "F": 12,   # Enabled

            # Reference
            "G": 20,   # Reference Frame

            # Properties
            "H": 14,   # Dynamic
            "I": 14,   # Respondable
            "J": 16,   # Visible
            "K": 14,   # Collidable
            "L": 14,   # Detectable
            "M": 14,   # Measurable

            # Position
            "N": 10,   # Pos X
            "O": 10,   # Pos Y
            "P": 10,   # Pos Z

            # Orientation
            "Q": 10,   # Rot X
            "R": 10,   # Rot Y
            "S": 10,   # Rot Z

            # Bounding Box
            "T": 12,   # BBox Min X
            "U": 12,   # BBox Min Y
            "V": 12,   # BBox Min Z

            "W": 12,   # BBox Size X
            "X": 12,   # BBox Size Y
            "Y": 12,   # BBox Size Z

            # Joint
            "Z": 18,   # Joint Type
            "AA": 10,  # Cyclic
            "AB": 14,  # Lower Limit
            "AC": 14,  # Upper Limit
            "AD": 12,  # Motor

            # Notes
            "AE": 40

        }

        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        # --------------------------------------------------
        # Protect Worksheet
        # --------------------------------------------------
        self.protect_worksheet(
            ws,
            editable_columns=[
                # Object
                "D",   # Object Name
                "E",   # Object Type
                "F",   # Enabled

                # Reference
                "G",   # Reference Frame

                # Properties
                "H",   # Dynamic
                "I",   # Respondable
                "J",   # Visible
                "K",   # Collidable
                "L",   # Detectable
                "M",   # Measurable

                # Position
                "N",   # Pos X
                "O",   # Pos Y
                "P",   # Pos Z

                # Orientation
                "Q",   # Rot X
                "R",   # Rot Y
                "S",   # Rot Z

                # Bounding Box
                "T",   # BBox Min X
                "U",   # BBox Min Y
                "V",   # BBox Min Z
                "W",   # BBox Size X
                "X",   # BBox Size Y
                "Y",   # BBox Size Z

                # Joint
                "Z",   # Joint Type
                "AA",  # Cyclic
                "AB",  # Lower Limit
                "AC",  # Upper Limit
                "AD",  # Motor

                # Notes
                "AE"
            ]
        )

        self.logger.info(
            "Objects worksheet created (%d objects).",
            row - 2
        )
 
    def create_lists_sheet(self, workbook):

        self.logger.info("Creating worksheet: Lists")

        ws = workbook.create_sheet("Lists")

        # ==================================================
        # Components
        # ==================================================

        ws["A1"] = "Component Name"

        row = 2

        for component in self.database["components"]:

            ws.cell(row=row, column=1).value = component["name"]
            row += 1

        last_component = row - 1

        # ==================================================
        # Boolean
        # ==================================================

        ws["B1"] = "Boolean"
        ws["B2"] = True
        ws["B3"] = False

        # ==================================================
        # Component Types
        # ==================================================

        ws["C1"] = "Component Type"

        values = [

            "Mechanical",
            "Electrical",
            "Electronic",
            "Sensor",
            "Actuator",
            "Structure",
            "Fastener",
            "Other"

        ]

        row = 2

        for value in values:

            ws.cell(row=row, column=3).value = value
            row += 1

        last_component_type = row - 1

        # ==================================================
        # Object Types
        # ==================================================

        ws["D1"] = "Object Type"

        values = [

            "Shape",
            "Joint",
            "Dummy",
            "Camera",
            "Light",
            "Vision Sensor",
            "Proximity Sensor",
            "Force Sensor",
            "Path",
            "Point Cloud",
            "Octree"

        ]

        row = 2

        for value in values:

            ws.cell(row=row, column=4).value = value
            row += 1

        last_object_type = row - 1

        # ==================================================
        # Reference Frame
        # ==================================================

        ws["E1"] = "Reference Frame"

        ws["E2"] = "parent"
        ws["E3"] = "world"

        # ==================================================
        # Joint Types
        # ==================================================

        ws["F1"] = "Joint Type"

        values = [

            "Revolute",
            "Prismatic",
            "Spherical"

        ]

        row = 2

        for value in values:

            ws.cell(row=row, column=6).value = value
            row += 1

        last_joint_type = row - 1

        # ==================================================
        # Status
        # ==================================================

        ws["G1"] = "Status"

        values = [

            "NEW",
            "IN_PROGRESS",
            "REVIEWED",
            "VERIFIED",
            "DONE"

        ]

        row = 2

        for value in values:

            ws.cell(row=row, column=7).value = value
            row += 1

        last_status = row - 1

        # ==================================================
        # Motor Mode
        # ==================================================

        ws["H1"] = "Motor Mode"

        values = [

            "Position",
            "Velocity",
            "Torque"

        ]

        row = 2

        for value in values:

            ws.cell(row=row, column=8).value = value
            row += 1

        last_motor = row - 1

        # ==================================================
        # Column Widths
        # ==================================================

        widths = {

            "A": 30,
            "B": 12,
            "C": 22,
            "D": 22,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 18

        }

        for col, width in widths.items():

            ws.column_dimensions[col].width = width

        # ==================================================
        # Named Ranges
        # ==================================================

        workbook.defined_names.add(
            DefinedName(
                "ComponentNames",
                attr_text=f"Lists!$A$2:$A${last_component}"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "Boolean",
                attr_text="Lists!$B$2:$B$3"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "ComponentType",
                attr_text=f"Lists!$C$2:$C${last_component_type}"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "ObjectType",
                attr_text=f"Lists!$D$2:$D${last_object_type}"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "ReferenceFrame",
                attr_text="Lists!$E$2:$E$3"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "JointType",
                attr_text=f"Lists!$F$2:$F${last_joint_type}"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "Status",
                attr_text=f"Lists!$G$2:$G${last_status}"
            )
        )

        workbook.defined_names.add(
            DefinedName(
                "MotorMode",
                attr_text=f"Lists!$H$2:$H${last_motor}"
            )
        )

        # ==================================================
        # Hide Worksheet
        # ==================================================

        ws.sheet_state = "veryHidden"

        self.logger.info("Lists worksheet created.")
 
    def apply_excel_styles(self, workbook):

        self.logger.info("Applying Excel styles")

        # --------------------------------------------------
        # Colors
        # --------------------------------------------------

        HEADER_FILL = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        AUTO_FILL = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE"
        )

        EDIT_FILL = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

        HEADER_FONT = Font(
            bold=True,
            color="FFFFFF"
        )

        THIN = Side(
            border_style="thin",
            color="A6A6A6"
        )

        BORDER = Border(
            left=THIN,
            right=THIN,
            top=THIN,
            bottom=THIN
        )

        CENTER = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ==================================================
        # Components
        # ==================================================

        ws = workbook["Components"]

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Automatic columns

        for col in ["A", "B", "E", "G"]:

            for cell in ws[col][1:]:

                cell.fill = AUTO_FILL
                cell.border = BORDER

        # Editable columns

        for col in ["C", "D", "F", "H", "I", "J", "K"]:

            for cell in ws[col][1:]:

                cell.fill = EDIT_FILL
                cell.border = BORDER

        # ==================================================
        # Objects
        # ==================================================

        ws = workbook["Objects"]

        for cell in ws[1]:

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

        # Automatically generated columns

        for col in ["A", "B", "C","D", "F"]:

            for cell in ws[col][1:]:

                cell.fill = AUTO_FILL
                cell.border = BORDER

        # Editable columns

        for col in [

            "E",  # Object Type
            "G",  # Reference Frame

            "H",
            "I",
            "J",
            "K",
            "L",
            "M",

            "N",
            "O",
            "P",

            "Q",
            "R",
            "S",

            "T",
            "U",
            "V",

            "W",
            "X",
            "Y",

            "Z",
            "AA",
            "AB",
            "AC",
            "AD",

            "AE"

        ]:

            for cell in ws[col][1:]:

                cell.fill = EDIT_FILL
                cell.border = BORDER

        # ==================================================
        # Center all cells
        # ==================================================

        for worksheet in workbook.worksheets:

            for row in worksheet.iter_rows():

                for cell in row:

                    cell.alignment = CENTER


        self.logger.info("Excel styles applied.")

    def apply_excel_validations(self, workbook):

        self.logger.info("Applying Excel validations")

        ws_components = workbook["Components"]
        ws_objects = workbook["Objects"]

        # ==================================================
        # Components
        # ==================================================

        dv_component_type = DataValidation(
            type="list",
            formula1="=ComponentType",
            allow_blank=True
        )

        dv_component_parent = DataValidation(
            type="list",
            formula1="=ComponentNames",
            allow_blank=True
        )

        dv_component_boolean = DataValidation(
            type="list",
            formula1="=Boolean",
            allow_blank=False
        )

        dv_component_status = DataValidation(
            type="list",
            formula1="=Status",
            allow_blank=False
        )

        ws_components.add_data_validation(dv_component_type)
        ws_components.add_data_validation(dv_component_parent)
        ws_components.add_data_validation(dv_component_boolean)
        ws_components.add_data_validation(dv_component_status)

        # ==================================================
        # Objects
        # ==================================================

        dv_object_type = DataValidation(
            type="list",
            formula1="=ObjectType",
            allow_blank=False
        )

        dv_reference_frame = DataValidation(
            type="list",
            formula1="=ReferenceFrame",
            allow_blank=False
        )

        dv_boolean = DataValidation(
            type="list",
            formula1="=Boolean",
            allow_blank=False
        )

        dv_joint_type = DataValidation(
            type="list",
            formula1="=JointType",
            allow_blank=True
        )

        dv_motor_mode = DataValidation(
            type="list",
            formula1="=MotorMode",
            allow_blank=True
        )

        ws_objects.add_data_validation(dv_object_type)
        ws_objects.add_data_validation(dv_reference_frame)
        ws_objects.add_data_validation(dv_boolean)
        ws_objects.add_data_validation(dv_joint_type)
        ws_objects.add_data_validation(dv_motor_mode)

        # ==================================================
        # Components Sheet
        # ==================================================

        for row in range(2, ws_components.max_row + 1):

            dv_component_type.add(ws_components[f"C{row}"])
            dv_component_parent.add(ws_components[f"F{row}"])
            dv_component_boolean.add(ws_components[f"H{row}"])
            dv_component_boolean.add(ws_components[f"I{row}"])
            dv_component_status.add(ws_components[f"J{row}"])

        # ==================================================
        # Objects Sheet
        # ==================================================

        for row in range(2, ws_objects.max_row + 1):

            # Object Type
            dv_object_type.add(ws_objects[f"E{row}"])

            # Enabled
            dv_boolean.add(ws_objects[f"F{row}"])

            # Reference Frame
            dv_reference_frame.add(ws_objects[f"G{row}"])

            # Dynamic
            dv_boolean.add(ws_objects[f"H{row}"])

            # Respondable
            dv_boolean.add(ws_objects[f"I{row}"])

            # Visible
            dv_boolean.add(ws_objects[f"J{row}"])

            # Collidable
            dv_boolean.add(ws_objects[f"K{row}"])

            # Detectable
            dv_boolean.add(ws_objects[f"L{row}"])

            # Measurable
            dv_boolean.add(ws_objects[f"M{row}"])

            # Joint Type
            dv_joint_type.add(ws_objects[f"Z{row}"])

            # Cyclic
            dv_boolean.add(ws_objects[f"AA{row}"])

            # Motor
            dv_motor_mode.add(ws_objects[f"AD{row}"])

        self.logger.info(
            "Excel validations successfully applied."
        )
    def save_assembly_table(self, workbook):

        self.logger.info("Saving Assembly Table")

        # --------------------------------------------------
        # Workbook Properties
        # --------------------------------------------------

        workbook.properties.creator = "GSPL-01_Rhino_Extractor"

        workbook.properties.title = "GSPL Assembly Table"

        workbook.properties.subject = "GIAR Simulation Pipeline for LiDAR"

        workbook.properties.description = (
            "Assembly definition table automatically generated "
            "by GSPL-01_Rhino_Extractor."
        )

        workbook.properties.company = "GIAR"

        workbook.properties.category = "Engineering"

        workbook.properties.keywords = (
            "GSPL, Rhino, CoppeliaSim, Assembly, LiDAR"
        )

        # --------------------------------------------------
        # Output File
        # --------------------------------------------------

        output_file = (

            Path(self.cfg["paths"]["database_directory"])

            /

            self.cfg["files"]["assembly_table"]

        )

        # --------------------------------------------------
        # Save Workbook
        # --------------------------------------------------

        try:

            workbook.save(output_file)

            self.logger.info("")
            self.logger.info("Assembly table successfully saved.")
            self.logger.info("File : %s", output_file)

        except Exception as e:

            self.logger.error("")
            self.logger.error("Unable to save assembly table.")
            self.logger.error(str(e))
            raise
    def generate_assembly_table(self):

        self.logger.info("")
        self.logger.info("========================================================")
        self.logger.info(" Generating Assembly Table")
        self.logger.info("========================================================")

        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)
        self.create_components_sheet(workbook)
        self.create_objects_sheet(workbook)
        self.create_lists_sheet(workbook)
        self.update_parent_ids(workbook)
        self.apply_excel_styles(workbook)
        self.apply_excel_validations(workbook)
        self.save_assembly_table(workbook)

        self.logger.info("Assembly table successfully generated.")
        self.logger.info("========================================================")

    # ------------------------------------------------------
    # Run
    # ------------------------------------------------------
    def run(self):

        self.load_config()
        self.configure_logger()
        self.print_banner()
        self.verify_project()
        self.open_model()
        self.audit_geometry()
        self.print_model_information()
        self.inspect_objects()
 
        self.build_database()
        self.build_geometry_database()
        self.save_database()
        self.generate_assembly_template()
        self.generate_assembly_table()
 
        self.logger.info("")
        self.logger.info("Finished.")



# ==========================================================
# Main
# ==========================================================
def main():

    extractor = RhinoExtractor()

    extractor.run()


if __name__ == "__main__":

    main()